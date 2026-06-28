import uuid
import io
import csv
from datetime import datetime, date, timezone
from typing import Any, Dict, List, Optional, Tuple, Union
from fastapi import HTTPException, status
from sqlalchemy import select, func, or_
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

import openpyxl

from app.db.session import get_db
from app.models.employee import Employee, Department, Designation, Branch, EmployeeStatus
from app.models.device import Device
from app.models.command import DeviceCommand, CommandType, CommandStatus

class EmployeeService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_employee(self, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Employee:
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        emp_code = data.get("employee_code")
        email = data.get("email")
        dev_user_id = data.get("device_user_id")

        if not emp_code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="employee_code is required")

        # Check unique constraints
        stmt_code = select(Employee).where(Employee.tenant_id == tenant_id, Employee.employee_code == emp_code)
        if (await self.db.execute(stmt_code)).scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Employee code already exists")

        if email:
            stmt_email = select(Employee).where(Employee.tenant_id == tenant_id, Employee.email == email)
            if (await self.db.execute(stmt_email)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Employee email already exists")

        if dev_user_id:
            stmt_dev = select(Employee).where(Employee.tenant_id == tenant_id, Employee.device_user_id == dev_user_id)
            if (await self.db.execute(stmt_dev)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Device user ID already exists")

        employee = Employee(tenant_id=tenant_id, **data)
        self.db.add(employee)
        await self.db.commit()
        await self.db.refresh(employee)
        return employee

    async def get_employee(self, employee_id: uuid.UUID, tenant_id: uuid.UUID) -> Employee:
        stmt = select(Employee).where(
            Employee.id == employee_id,
            Employee.tenant_id == tenant_id
        ).options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
            selectinload(Employee.branch),
            selectinload(Employee.shift)
        )
        res = await self.db.execute(stmt)
        employee = res.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
        return employee

    async def get_employee_by_code(self, employee_code: str, tenant_id: uuid.UUID) -> Employee:
        stmt = select(Employee).where(
            Employee.employee_code == employee_code,
            Employee.tenant_id == tenant_id
        ).options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
            selectinload(Employee.branch),
            selectinload(Employee.shift)
        )
        res = await self.db.execute(stmt)
        employee = res.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")
        return employee

    async def update_employee(self, employee_id: uuid.UUID, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Employee:
        employee = await self.get_employee(employee_id, tenant_id)
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        emp_code = data.get("employee_code")
        email = data.get("email")
        dev_user_id = data.get("device_user_id")

        if emp_code and emp_code != employee.employee_code:
            stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.employee_code == emp_code)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Employee code already exists")

        if email and email != employee.email:
            stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.email == email)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Employee email already exists")

        if dev_user_id and dev_user_id != employee.device_user_id:
            stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.device_user_id == dev_user_id)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Device user ID already exists")

        for field, val in data.items():
            if hasattr(employee, field):
                setattr(employee, field, val)

        await self.db.commit()
        await self.db.refresh(employee)
        return employee

    async def delete_employee(self, employee_id: uuid.UUID, tenant_id: uuid.UUID) -> None:
        employee = await self.get_employee(employee_id, tenant_id)
        await self.db.delete(employee)
        await self.db.commit()

    async def list_employees(
        self,
        tenant_id: uuid.UUID,
        department_id: Optional[uuid.UUID] = None,
        designation_id: Optional[uuid.UUID] = None,
        branch_id: Optional[uuid.UUID] = None,
        status: Optional[str] = None,
        search: Optional[str] = None,
        page: int = 1,
        page_size: int = 20
    ) -> Tuple[List[Employee], int]:
        count_stmt = select(func.count(Employee.id)).where(Employee.tenant_id == tenant_id)
        stmt = select(Employee).where(Employee.tenant_id == tenant_id).options(
            selectinload(Employee.department),
            selectinload(Employee.designation),
            selectinload(Employee.branch),
            selectinload(Employee.shift)
        )

        if department_id:
            count_stmt = count_stmt.where(Employee.department_id == department_id)
            stmt = stmt.where(Employee.department_id == department_id)
        if designation_id:
            count_stmt = count_stmt.where(Employee.designation_id == designation_id)
            stmt = stmt.where(Employee.designation_id == designation_id)
        if branch_id:
            count_stmt = count_stmt.where(Employee.branch_id == branch_id)
            stmt = stmt.where(Employee.branch_id == branch_id)
        if status:
            count_stmt = count_stmt.where(Employee.status == status)
            stmt = stmt.where(Employee.status == status)
        if search:
            escaped = search.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            search_filter = or_(
                Employee.first_name.ilike(f"%{escaped}%"),
                Employee.last_name.ilike(f"%{escaped}%"),
                Employee.employee_code.ilike(f"%{escaped}%"),
                Employee.email.ilike(f"%{escaped}%")
            )
            count_stmt = count_stmt.where(search_filter)
            stmt = stmt.where(search_filter)

        total_res = await self.db.execute(count_stmt)
        total = total_res.scalar() or 0

        stmt = stmt.offset((page - 1) * page_size).limit(page_size)
        res = await self.db.execute(stmt)
        employees = list(res.scalars().all())
        return employees, total

    async def sync_to_device(self, employee_id: uuid.UUID, tenant_id: uuid.UUID) -> List[DeviceCommand]:
        employee = await self.get_employee(employee_id, tenant_id)
        if not employee.branch_id:
            return []

        # Find devices in this branch
        stmt = select(Device).where(Device.branch_id == employee.branch_id, Device.tenant_id == tenant_id)
        res = await self.db.execute(stmt)
        devices = res.scalars().all()

        commands = []
        for device in devices:
            cmd = DeviceCommand(
                tenant_id=tenant_id,
                device_id=device.id,
                command_type=CommandType.ENROLL_FP,  # standard enroll/sync command
                parameters={
                    "employee_id": str(employee.id),
                    "employee_code": employee.employee_code,
                    "first_name": employee.first_name,
                    "last_name": employee.last_name,
                    "device_user_id": employee.device_user_id
                },
                status=CommandStatus.PENDING,
                requested_at=datetime.now(timezone.utc)
            )
            self.db.add(cmd)
            commands.append(cmd)

        if commands:
            await self.db.commit()
            for cmd in commands:
                await self.db.refresh(cmd)
        return commands

    async def bulk_import(self, tenant_id: uuid.UUID, file_content: bytes, filename: str = "import.xlsx") -> Tuple[int, List[str]]:
        errors = []
        rows_to_process = []
        is_excel = file_content.startswith(b"PK\x03\x04")

        if is_excel:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
                sheet = wb.active
                headers = []
                for cell in sheet[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip().lower().replace(" ", "_"))
                    else:
                        headers.append("")

                for r in range(2, sheet.max_row + 1):
                    row_data = {}
                    has_data = False
                    for col_idx, header in enumerate(headers):
                        if header:
                            val = sheet.cell(row=r, column=col_idx + 1).value
                            if val is not None:
                                has_data = True
                                row_data[header] = str(val).strip()
                            else:
                                row_data[header] = None
                    if has_data:
                        rows_to_process.append((r, row_data))
            except Exception as e:
                return 0, [f"Failed to parse Excel file: {str(e)}"]
        else:
            try:
                try:
                    text = file_content.decode("utf-8")
                except UnicodeDecodeError:
                    text = file_content.decode("latin-1")
                f = io.StringIO(text)
                reader = csv.reader(f)
                try:
                    headers = [h.strip().lower().replace(" ", "_") for h in next(reader)]
                except StopIteration:
                    headers = []

                for r_idx, row in enumerate(reader, start=2):
                    if not row or not any(row):
                        continue
                    row_data = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            val = row[col_idx].strip()
                            row_data[header] = val if val else None
                    rows_to_process.append((r_idx, row_data))
            except Exception as e:
                return 0, [f"Failed to parse CSV file: {str(e)}"]

        if not rows_to_process:
            return 0, ["No data found in the file"]

        # Cache departments, designations, branches
        dept_res = await self.db.execute(select(Department).where(Department.tenant_id == tenant_id))
        depts = {d.code.lower(): d.id for d in dept_res.scalars().all() if d.code}

        desg_res = await self.db.execute(select(Designation).where(Designation.tenant_id == tenant_id))
        desgs = {d.code.lower(): d.id for d in desg_res.scalars().all() if d.code}

        branch_res = await self.db.execute(select(Branch).where(Branch.tenant_id == tenant_id))
        branches = {b.code.lower(): b.id for b in branch_res.scalars().all() if b.code}

        created_count = 0

        for row_num, data in rows_to_process:
            emp_code = data.get("employee_code")
            first_name = data.get("first_name")
            last_name = data.get("last_name")

            if not emp_code:
                errors.append(f"Row {row_num}: employee_code is required")
                continue
            if not first_name:
                errors.append(f"Row {row_num}: first_name is required")
                continue
            if not last_name:
                errors.append(f"Row {row_num}: last_name is required")
                continue

            # Validate unique constraints
            stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.employee_code == emp_code)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                errors.append(f"Row {row_num}: employee_code '{emp_code}' already exists")
                continue

            email = data.get("email")
            if email:
                stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.email == email)
                if (await self.db.execute(stmt)).scalar_one_or_none():
                    errors.append(f"Row {row_num}: email '{email}' already exists")
                    continue

            dev_user_id = data.get("device_user_id")
            if dev_user_id:
                stmt = select(Employee).where(Employee.tenant_id == tenant_id, Employee.device_user_id == dev_user_id)
                if (await self.db.execute(stmt)).scalar_one_or_none():
                    errors.append(f"Row {row_num}: device_user_id '{dev_user_id}' already exists")
                    continue

            dept_id = None
            dept_code = data.get("department_code")
            if dept_code:
                dept_id = depts.get(dept_code.lower())
                if not dept_id:
                    errors.append(f"Row {row_num}: department_code '{dept_code}' not found")
                    continue

            desg_id = None
            desg_code = data.get("designation_code")
            if desg_code:
                desg_id = desgs.get(desg_code.lower())
                if not desg_id:
                    errors.append(f"Row {row_num}: designation_code '{desg_code}' not found")
                    continue

            br_id = None
            br_code = data.get("branch_code")
            if br_code:
                br_id = branches.get(br_code.lower())
                if not br_id:
                    errors.append(f"Row {row_num}: branch_code '{br_code}' not found")
                    continue

            joining_date = None
            joining_date_str = data.get("joining_date")
            if joining_date_str:
                try:
                    joining_date = datetime.strptime(joining_date_str, "%Y-%m-%d").date()
                except ValueError:
                    try:
                        joining_date = datetime.strptime(joining_date_str, "%d/%m/%Y").date()
                    except ValueError:
                        errors.append(f"Row {row_num}: invalid joining_date format (expected YYYY-MM-DD)")
                        continue

            dob = None
            dob_str = data.get("date_of_birth")
            if dob_str:
                try:
                    dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
                except ValueError:
                    try:
                        dob = datetime.strptime(dob_str, "%d/%m/%Y").date()
                    except ValueError:
                        errors.append(f"Row {row_num}: invalid date_of_birth format (expected YYYY-MM-DD)")
                        continue

            status_val = data.get("status", "active").lower()
            if status_val not in [e.value for e in EmployeeStatus]:
                status_val = "active"

            employee = Employee(
                tenant_id=tenant_id,
                employee_code=emp_code,
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=data.get("phone"),
                photo_url=data.get("photo_url"),
                department_id=dept_id,
                designation_id=desg_id,
                branch_id=br_id,
                joining_date=joining_date,
                date_of_birth=dob,
                gender=data.get("gender"),
                address=data.get("address"),
                city=data.get("city"),
                state=data.get("state"),
                pincode=data.get("pincode"),
                emergency_contact_name=data.get("emergency_contact_name"),
                emergency_contact_phone=data.get("emergency_contact_phone"),
                blood_group=data.get("blood_group"),
                status=status_val,
                device_user_id=dev_user_id
            )
            self.db.add(employee)
            created_count += 1

        if created_count > 0:
            await self.db.commit()

        return created_count, errors


class DepartmentService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_department(self, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Department:
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if not code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="code is required")

        stmt = select(Department).where(Department.tenant_id == tenant_id, Department.code == code)
        if (await self.db.execute(stmt)).scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Department with this code already exists")

        dept = Department(tenant_id=tenant_id, **data)
        self.db.add(dept)
        await self.db.commit()
        await self.db.refresh(dept)
        return dept

    async def get_department(self, department_id: uuid.UUID, tenant_id: uuid.UUID) -> Department:
        stmt = select(Department).where(Department.id == department_id, Department.tenant_id == tenant_id)
        res = await self.db.execute(stmt)
        dept = res.scalar_one_or_none()
        if not dept:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Department not found")
        return dept

    async def update_department(self, department_id: uuid.UUID, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Department:
        dept = await self.get_department(department_id, tenant_id)
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if code and code != dept.code:
            stmt = select(Department).where(Department.tenant_id == tenant_id, Department.code == code)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Department with this code already exists")

        for field, val in data.items():
            if hasattr(dept, field):
                setattr(dept, field, val)

        await self.db.commit()
        await self.db.refresh(dept)
        return dept

    async def delete_department(self, department_id: uuid.UUID, tenant_id: uuid.UUID) -> None:
        dept = await self.get_department(department_id, tenant_id)
        await self.db.delete(dept)
        await self.db.commit()

    async def list_departments(self, tenant_id: uuid.UUID, page: int = 1, page_size: int = 20) -> Tuple[List[Department], int]:
        count_stmt = select(func.count(Department.id)).where(Department.tenant_id == tenant_id)
        stmt = select(Department).where(Department.tenant_id == tenant_id).offset((page - 1) * page_size).limit(page_size)
        
        total_res = await self.db.execute(count_stmt)
        total = total_res.scalar() or 0
        
        res = await self.db.execute(stmt)
        depts = list(res.scalars().all())
        return depts, total


class DesignationService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_designation(self, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Designation:
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if not code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="code is required")

        stmt = select(Designation).where(Designation.tenant_id == tenant_id, Designation.code == code)
        if (await self.db.execute(stmt)).scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Designation with this code already exists")

        desg = Designation(tenant_id=tenant_id, **data)
        self.db.add(desg)
        await self.db.commit()
        await self.db.refresh(desg)
        return desg

    async def get_designation(self, designation_id: uuid.UUID, tenant_id: uuid.UUID) -> Designation:
        stmt = select(Designation).where(Designation.id == designation_id, Designation.tenant_id == tenant_id)
        res = await self.db.execute(stmt)
        desg = res.scalar_one_or_none()
        if not desg:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Designation not found")
        return desg

    async def update_designation(self, designation_id: uuid.UUID, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Designation:
        desg = await self.get_designation(designation_id, tenant_id)
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if code and code != desg.code:
            stmt = select(Designation).where(Designation.tenant_id == tenant_id, Designation.code == code)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Designation with this code already exists")

        for field, val in data.items():
            if hasattr(desg, field):
                setattr(desg, field, val)

        await self.db.commit()
        await self.db.refresh(desg)
        return desg

    async def delete_designation(self, designation_id: uuid.UUID, tenant_id: uuid.UUID) -> None:
        desg = await self.get_designation(designation_id, tenant_id)
        await self.db.delete(desg)
        await self.db.commit()

    async def list_designations(self, tenant_id: uuid.UUID, page: int = 1, page_size: int = 20) -> Tuple[List[Designation], int]:
        count_stmt = select(func.count(Designation.id)).where(Designation.tenant_id == tenant_id)
        stmt = select(Designation).where(Designation.tenant_id == tenant_id).offset((page - 1) * page_size).limit(page_size)
        
        total_res = await self.db.execute(count_stmt)
        total = total_res.scalar() or 0
        
        res = await self.db.execute(stmt)
        desgs = list(res.scalars().all())
        return desgs, total


class BranchService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_branch(self, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Branch:
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if not code:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="code is required")

        stmt = select(Branch).where(Branch.tenant_id == tenant_id, Branch.code == code)
        if (await self.db.execute(stmt)).scalar_one_or_none():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Branch with this code already exists")

        branch = Branch(tenant_id=tenant_id, **data)
        self.db.add(branch)
        await self.db.commit()
        await self.db.refresh(branch)
        return branch

    async def get_branch(self, branch_id: uuid.UUID, tenant_id: uuid.UUID) -> Branch:
        stmt = select(Branch).where(Branch.id == branch_id, Branch.tenant_id == tenant_id)
        res = await self.db.execute(stmt)
        branch = res.scalar_one_or_none()
        if not branch:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Branch not found")
        return branch

    async def update_branch(self, branch_id: uuid.UUID, tenant_id: uuid.UUID, data: Union[Dict[str, Any], Any]) -> Branch:
        branch = await self.get_branch(branch_id, tenant_id)
        if not isinstance(data, dict):
            data = data.model_dump(exclude_unset=True)

        code = data.get("code")
        if code and code != branch.code:
            stmt = select(Branch).where(Branch.tenant_id == tenant_id, Branch.code == code)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Branch with this code already exists")

        for field, val in data.items():
            if hasattr(branch, field):
                setattr(branch, field, val)

        await self.db.commit()
        await self.db.refresh(branch)
        return branch

    async def delete_branch(self, branch_id: uuid.UUID, tenant_id: uuid.UUID) -> None:
        branch = await self.get_branch(branch_id, tenant_id)
        await self.db.delete(branch)
        await self.db.commit()

    async def list_branches(self, tenant_id: uuid.UUID, page: int = 1, page_size: int = 20) -> Tuple[List[Branch], int]:
        count_stmt = select(func.count(Branch.id)).where(Branch.tenant_id == tenant_id)
        stmt = select(Branch).where(Branch.tenant_id == tenant_id).offset((page - 1) * page_size).limit(page_size)
        
        total_res = await self.db.execute(count_stmt)
        total = total_res.scalar() or 0
        
        res = await self.db.execute(stmt)
        branches = list(res.scalars().all())
        return branches, total
