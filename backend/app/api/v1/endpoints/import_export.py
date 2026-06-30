"""Import & Export API endpoints."""

import uuid
import csv
import io
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_db, get_current_active_user, require_permissions
from app.models.user import User
from app.models.employee import Employee, Department, Designation, Branch
from app.models.leave import LeaveBalance, LeaveType
from app.middleware.rate_limit import rate_limit
import structlog

logger = structlog.get_logger(__name__)
router = APIRouter(dependencies=[Depends(require_permissions("employee.read"))])


@router.post("/import/employees", dependencies=[Depends(require_permissions("employee.create"))])
@rate_limit(limit=10, period=60)
async def import_employees(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Import employees from CSV/Excel file."""
    if not file.filename.endswith(('.csv', '.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")

    content = await file.read()

    try:
        if file.filename.endswith('.csv'):
            text = content.decode('utf-8')
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        elif file.filename.endswith(('.xlsx', '.xls')):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: str(row[i]).strip() if row[i] is not None else '' for i in range(len(headers))})
        else:
            raise HTTPException(status_code=400, detail="Only CSV and Excel files are supported")
    except Exception as e:
        logger.error("import_parse_failed", error=str(e), filename=file.filename)
        raise HTTPException(status_code=400, detail="Failed to parse file. Ensure it is valid CSV or Excel.")

    results = {"total": len(rows), "created": 0, "skipped": 0, "errors": []}

    for i, row in enumerate(rows):
        try:
            code = row.get('employee_code', '').strip()
            first_name = row.get('first_name', '').strip()
            last_name = row.get('last_name', '').strip()

            if not code or not first_name:
                results["errors"].append(f"Row {i+1}: Missing employee_code or first_name")
                results["skipped"] += 1
                continue

            existing = await db.execute(
                select(Employee).where(
                    Employee.tenant_id == current_user.tenant_id,
                    Employee.employee_code == code,
                )
            )
            if existing.scalar_one_or_none():
                results["skipped"] += 1
                continue

            # Resolve department
            dept_id = None
            dept_name = row.get('department', '').strip()
            if dept_name:
                dept = await db.execute(
                    select(Department).where(
                        Department.tenant_id == current_user.tenant_id,
                        Department.name == dept_name,
                    )
                )
                dept_obj = dept.scalar_one_or_none()
                if dept_obj:
                    dept_id = dept_obj.id

            # Resolve designation
            desig_id = None
            desig_name = row.get('designation', '').strip()
            if desig_name:
                desig = await db.execute(
                    select(Designation).where(
                        Designation.tenant_id == current_user.tenant_id,
                        Designation.name == desig_name,
                    )
                )
                desig_obj = desig.scalar_one_or_none()
                if desig_obj:
                    desig_id = desig_obj.id

            # Resolve branch
            branch_id = None
            branch_name = row.get('branch', '').strip()
            if branch_name:
                branch = await db.execute(
                    select(Branch).where(
                        Branch.tenant_id == current_user.tenant_id,
                        Branch.name == branch_name,
                    )
                )
                branch_obj = branch.scalar_one_or_none()
                if branch_obj:
                    branch_id = branch_obj.id

            employee = Employee(
                tenant_id=current_user.tenant_id,
                employee_code=code,
                first_name=first_name,
                last_name=last_name,
                email=row.get('email', '').strip() or None,
                phone=row.get('phone', '').strip() or None,
                gender=row.get('gender', '').strip().lower() or None,
                department_id=dept_id,
                designation_id=desig_id,
                branch_id=branch_id,
                status="active",
            )
            db.add(employee)
            results["created"] += 1

        except Exception as e:
            logger.error("employee_import_row_failed", row=i+1, error=str(e))
            results["errors"].append(f"Row {i+1}: Failed to import employee")
            results["skipped"] += 1

    await db.commit()
    return results


@router.post("/import/leave-balances", dependencies=[Depends(require_permissions("employee.manage"))])
@rate_limit(limit=10, period=60)
async def import_leave_balances(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Import leave balances from CSV."""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files supported")

    content = await file.read()
    text = content.decode('utf-8')
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    results = {"total": len(rows), "created": 0, "updated": 0, "errors": []}

    for i, row in enumerate(rows):
        try:
            code = row.get('employee_code', '').strip()
            leave_type_name = row.get('leave_type', '').strip()
            balance = float(row.get('balance', 0))

            if not code or not leave_type_name:
                results["errors"].append(f"Row {i+1}: Missing employee_code or leave_type")
                continue

            emp = await db.execute(
                select(Employee).where(
                    Employee.tenant_id == current_user.tenant_id,
                    Employee.employee_code == code,
                )
            )
            employee = emp.scalar_one_or_none()
            if not employee:
                results["errors"].append(f"Row {i+1}: Employee {code} not found")
                continue

            lt = await db.execute(
                select(LeaveType).where(
                    LeaveType.tenant_id == current_user.tenant_id,
                    LeaveType.name == leave_type_name,
                )
            )
            leave_type = lt.scalar_one_or_none()
            if not leave_type:
                results["errors"].append(f"Row {i+1}: Leave type {leave_type_name} not found")
                continue

            existing = await db.execute(
                select(LeaveBalance).where(
                    LeaveBalance.tenant_id == current_user.tenant_id,
                    LeaveBalance.employee_id == employee.id,
                    LeaveBalance.leave_type_id == leave_type.id,
                )
            )
            lb = existing.scalar_one_or_none()
            if lb:
                lb.total_days = balance
                results["updated"] += 1
            else:
                db.add(LeaveBalance(
                    tenant_id=current_user.tenant_id,
                    employee_id=employee.id,
                    leave_type_id=leave_type.id,
                    year=date.today().year,
                    total_days=balance,
                    used_days=0,
                ))
                results["created"] += 1

        except Exception as e:
            logger.error("leave_balance_import_row_failed", row=i+1, error=str(e))
            results["errors"].append(f"Row {i+1}: Failed to import leave balance")

    await db.commit()
    return results


@router.get("/export/employees")
async def export_employees(
    format: str = Query("csv"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Export employees to CSV."""
    stmt = select(Employee).where(
        Employee.tenant_id == current_user.tenant_id
    ).order_by(Employee.employee_code)
    result = await db.execute(stmt)
    employees = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'employee_code', 'first_name', 'last_name', 'email', 'phone',
        'gender', 'status', 'joining_date', 'created_at',
    ])

    for emp in employees:
        writer.writerow([
            emp.employee_code, emp.first_name, emp.last_name,
            emp.email or '', emp.phone or '', emp.gender or '',
            emp.status, str(emp.joining_date) if emp.joining_date else '',
            emp.created_at.isoformat() if emp.created_at else '',
        ])

    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=employees.csv"},
    )


@router.get("/template/employees")
async def employee_import_template():
    """Download employee import template."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'employee_code', 'first_name', 'last_name', 'email', 'phone',
        'gender', 'department', 'designation', 'branch', 'joining_date',
    ])
    writer.writerow(['EMP001', 'John', 'Doe', 'john@example.com', '9876543210',
                      'male', 'Engineering', 'Developer', 'Head Office', '2024-01-15'])

    from fastapi.responses import Response
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=employee_import_template.csv"},
    )
