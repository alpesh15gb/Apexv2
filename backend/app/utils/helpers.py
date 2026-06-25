"""Utility functions for file export, pagination, and common operations."""

import csv
import io
import uuid
from datetime import datetime, date
from typing import Any, Optional

from sqlalchemy import select, func, desc, asc
from sqlalchemy.ext.asyncio import AsyncSession


def paginate_query(stmt, page: int = 1, page_size: int = 20):
    """Apply pagination to a SQLAlchemy select statement."""
    offset = (page - 1) * page_size
    return stmt.offset(offset).limit(page_size)


def generate_pass_number() -> str:
    """Generate a unique visitor pass number."""
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    random_suffix = uuid.uuid4().hex[:6].upper()
    return f"VP-{timestamp}-{random_suffix}"


def generate_badge_number() -> str:
    """Generate a unique badge number."""
    return f"B-{uuid.uuid4().hex[:8].upper()}"


def csv_to_dicts(content: bytes, filename: str) -> list[dict[str, Any]]:
    """Parse CSV or Excel file content into list of dicts."""
    if filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower().replace(" ", "_") for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(row)]
    else:
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader)


def dicts_to_csv(data: list[dict], fieldnames: list[str]) -> bytes:
    """Convert list of dicts to CSV bytes."""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue().encode("utf-8")


def dicts_to_excel(data: list[dict], fieldnames: list[str], sheet_name: str = "Sheet1") -> bytes:
    """Convert list of dicts to Excel bytes."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(fieldnames)
    for row in data:
        ws.append([row.get(f, "") for f in fieldnames])
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def calculate_business_days(start_date: date, end_date: date) -> int:
    """Calculate number of business days between two dates."""
    from datetime import timedelta
    days = 0
    current = start_date
    while current <= end_date:
        if current.weekday() < 5:
            days += 1
        current += timedelta(days=1)
    return days
